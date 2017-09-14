#!/usr/bin/env python
#Skrypt składający w jeden zbiorczy plik inormacje z mediaplanów w bieżącym katalogu

import os
import re
import time
from datetime import timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, range_boundaries, get_column_interval
import pandas as pd
from pandas import DataFrame

def listMpFiles():
    """Tworzy listę plików z poszczególnymi mediaplanami w bieżącym katalogu"""
    mp_files = [f for f in os.listdir('.') if re.match(r'\b[a-zA-Z0-9].*.xls*', f)]
    num_files = len(mp_files)
    print('Znalazłem {} plików.'.format(num_files))
    return mp_files

def sheetToDataFrame(range_string, ws, kolumna):
    """Tworzy DataFrame z zakresu komórek 'A4':podana ostatnia kolumna i wiersz danych"""
    adres = range_boundaries(range_string)

    data_rows = []

    for row in ws.iter_rows(min_col=adres[0], min_row=adres[1], max_col=adres[2], max_row=adres[3]):
        data_rows.append([cell.value for cell in row])
    # TODO usunąć puste wiersze z DataFrame
    working_df = DataFrame(data_rows)

    return working_df

#if __name__ == __main__:
start_time = time.monotonic()
mediaplany = listMpFiles()
baza = Workbook()
arkuszBazy = baza.active
currentSheetDF = DataFrame()
finalDF = DataFrame()
exportDF = DataFrame()
header_lv1 = []
header_lv2 = []

arkusze = ('LIC - MP', 'LIC - H', 'SUM - MP', 'SUM - H', 'SP - MP', 'SP - H', 'MBA - MP', 'MBA - H',
           'Szkolenia - MP', 'Ogólne',)

# pętla ładująca poszczególne pliki z mediaplanami i sprawdzająca czy zawierają potrzebne arkusze
for filename in mediaplany:
    try:
        wb = load_workbook(filename, keep_vba=True, data_only=True)
    except:
        print('Błąd otwarcia pliku: {}'.format(filename))

    print('Sprawdzam plik: {}'.format(filename))

    for arkusz in arkusze: #ładuje poszczególne arkusze i zbiera z nich dane
        try:
            ws = wb[arkusz]
        except:
            print('\t! -> Nie znalazłem arkusza {} w pliku {}, pomijam.'.format(arkusz, filename))
            continue

        lastRow = 0

        #Ustala ostatni wiersz danych
        searchCol = ws['A']
        sumaInRow = 0
        for searchCell in searchCol:
            sumaInRow += 1
            if searchCell.value == 'SUMA':
                lastRow = sumaInRow - 1
                # print('\tOstatni wiersz danych w {} to {}'.format(arkusz, lastRow))
            else: continue

        #Ustala ostatnią kolumnę danych
        searchRow = ws[3]
        sumaInCol = 0
        for searchCell in searchRow:
            sumaInCol += 1
            if searchCell.value != 'ga:source':
                continue
            else:
                kolumna = get_column_letter(sumaInCol)
                koniec = kolumna + str(lastRow)
                print('\t{} - {}: Wczytuję dane z zakresu [{}:{}]'.format(filename, arkusz, 'A4', koniec))

        final_data_cell = str(kolumna) + str(lastRow)
        cellsRange = 'A4' + ':' + str(final_data_cell)
        currentSheetDF = sheetToDataFrame(cellsRange, ws, kolumna)
        finalDF = finalDF.append(currentSheetDF)
    header_lv1 = [ws.cell(row=2, column=i).value for i in range(1, sumaInCol + 1)] # nagłówek wyższego poziomu
    header_lv2 = [ws.cell(row=3, column=i).value for i in range(1, sumaInCol + 1)] # nagłówek niższego poziomu
    header = [header_lv1, header_lv2]
    print('Plik: {} OK!'.format(filename))

# TODO ukryć kolumny, które były ukryte w pliku źródłowym
'''
przydatne linki:
https://stackoverflow.com/questions/38527725/how-can-i-hide-columns-in-openpyxl
https://stackoverflow.com/questions/31257353/finding-hidden-cells-using-openpyxl
'''
# dodaje najpierw wiersze nagłówka
exportDF = exportDF.append(header, ignore_index=True)
# dodaje wiersze danych
exportDF = exportDF.append(finalDF)

writer = pd.ExcelWriter('baza.xlsx', engine='xlsxwriter')

exportDF.to_excel(writer,
                 sheet_name='Baza',
                 index=False)

workbook = writer.book
worksheet = writer.sheets['Baza']
writer.save()
end_time = time.monotonic()
print('Gotowe w: ', timedelta(seconds=end_time - start_time))
time.sleep(5)